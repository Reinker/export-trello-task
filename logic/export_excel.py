import logic.trello_api
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import PatternFill, Border, Side
import json
import csv
import os

class ExportExcel:
    def __init__(self, json_file):
        self.__json_file = json_file
        self.__contents = {}
        self.__tags = ['フェーズ', 'ID', 'タスク名', 'タスク説明', '開始日', '更新日', 'ステータス', '担当者']
        self.__wb = Workbook()
        self.__ws = self.__wb.active
        self.__ws.title = 'タスク一覧'
        self.__fill = PatternFill(fill_type='solid', fgColor='00FF00')
        side = Side(style='thin', color='000000')
        self.__border = Border(top=side,bottom=side,right=side,left=side)

    def import_from_files(self):
        files = os.listdir('./jsons')
        if files.count == 0 :
            return

        for f in files:
            print('load file : ' + f)
            file_open = open('./jsons/' + f, 'r')
            content = json.load(file_open)
            api = logic.trello_api.TrelloAPI(content)
            board = api.get_board()
            api.sort_cards_by_date()
            self.__contents[board.get_name()] = board.get_cards()

    def __setTags(self):
        self.__ws.column_dimensions['A'].width = 20
        self.__ws.column_dimensions['B'].width = 30
        self.__ws.column_dimensions['C'].width = 10
        self.__ws.column_dimensions['D'].width = 30
        self.__ws.column_dimensions['E'].width = 25
        self.__ws.column_dimensions['F'].width = 25
        for row in self.__ws.iter_rows(min_row=1, max_col=len(self.__tags), max_row=1):
            for cell in row:
                cell.fill = self.__fill
                cell.border = self.__border
                cell.value = self.__tags[cell.column - 1]

    def __set_performance(self):
        for key in self.__contents:
            for card in self.__contents[key]:
                print(card.get_date())


    def exportAsExcel(self):
        if self.__contents == {}:
            return

        self.__setTags()
        offset = 2
        for key in self.__contents:
            cards = self.__contents[key]
            self.__ws.cell(row=offset, column=1).value = key
            self.__ws.merge_cells(start_row=offset, start_column=1, end_row=offset, end_column=len(self.__tags))
            for row in range(offset + 1, len(cards) + 2):
                content = []
                content.append(cards[row - 3].get_card_id())
                content.append(cards[row - 3].get_name())
                content.append(cards[row - 3].get_desc())
                content.append(cards[row - 3].get_date())
                content.append(cards[row - 3].get_date_last_activity())
                content.append(cards[row - 3].get_listname())
                content.append(','.join(cards[row - 2].get_membernames()))
                for col in range(2, len(self.__tags) + 1):
                    self.__ws.cell(row=row, column=col).value = content[col - 2]
            offset += len(cards)
        self.__wb.save('test.xlsx')
