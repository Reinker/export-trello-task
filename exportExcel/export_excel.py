import trelloAPI.trello_api as trello_api
import trelloAPI.card as trello_card
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill, Border, Side
from datetime import datetime
from datetime import timezone
from datetime import date
import numpy as np
import calendar
import json
import csv
import os
import re

HEADER_ROW_START=1
HEADER_COL_START=1
ROW_START=5
DATA_ROW_START=7
NORMAL_SIDE = Side(style='thin', color='000000')
BORDER = Border(top=NORMAL_SIDE,bottom=NORMAL_SIDE,right=NORMAL_SIDE,left=NORMAL_SIDE)
CONVEX_BORDER = Border(top=NORMAL_SIDE,right=NORMAL_SIDE,left=NORMAL_SIDE)
CONVEX_DOWNWARD_BORDER = Border(bottom=NORMAL_SIDE,right=NORMAL_SIDE,left=NORMAL_SIDE)
TOP_FILL = PatternFill(fill_type='solid', fgColor='55FF55')
PHASE_FILL = PatternFill(fill_type='solid', fgColor='FFFF55')
DATE_FILL = PatternFill(fill_type='solid', fgColor='6666FF')

class ExportExcel:

    def __init__(self):
        self.__boards = []
        self.__wb = Workbook()
        self.__ws = self.__wb.active
        self.__ws.title = 'タスク一覧'
        self.__col_offset = 1
        self.__import_from_files()

    def __import_from_files(self):
        files = []
        try:
            files = os.listdir('./jsons')
        except FileNotFoundError:
            print('Create directory "jsons" and put json files')

        if files.count == 0 :
            return

        for f in files:
            if re.match(r'.*\.json', f) == None:
                continue
            print('load file : ' + f)
            file_open = open('./jsons/' + f, 'r')
            json_str = json.load(file_open)
            api = trello_api.TrelloAPI(json_str)
            api.map_to_board()
            api.sort_cards_by_date()
            self.__boards.append(api.board())

    def __set_item_name_cell(self, width, name):
        col_num = self.__col_offset 
        if width > 0:
            self.__ws.column_dimensions[get_column_letter(col_num)].width = width

        self.__ws.cell(ROW_START, col_num).value = name
        self.__ws.cell(row=ROW_START, column=col_num).fill = TOP_FILL
        self.__ws.cell(row=ROW_START+1, column=col_num).fill = TOP_FILL
        self.__ws.cell(row=ROW_START, column=col_num).border = CONVEX_BORDER
        self.__ws.cell(row=ROW_START+1, column=col_num).border = CONVEX_DOWNWARD_BORDER

        self.__col_offset += 1

    def __set_performance_date_cell(self, max_col):
        self.__set_item_name_cell(0, '実績')

        project_start_date = trello_api.get_project_start_date(self.__boards)
        dates = []
        for i in range(0, 3):
            year = project_start_date.year
            month = project_start_date.month + i
            try:
                dates_sum = sum(calendar.Calendar().monthdatescalendar(year, month), [])
                filtered_dates = list(filter(lambda x: x.month == month, dates_sum))
                if month == project_start_date.month:
                    filtered_dates = list(filter(lambda x: x.day >= project_start_date.day, filtered_dates))
                dates.append(filtered_dates)
            except calendar.IllegalMonthError:
                month = month - 12
                year += 1
                dates_sum = sum(calendar.Calendar().monthdatescalendar(year, month), [])
                filtered_dates = list(filter(lambda x: x.month == month, dates_sum))
                dates.append(filtered_dates)

        
        col_offset = self.__ws.max_column
        for date in dates:
            self.__ws.cell(ROW_START+1, col_offset).value = date[0].month
            for i in range(0, len(date)):
                self.__ws.cell(ROW_START, col_offset + i).fill = TOP_FILL
                self.__ws.cell(ROW_START+1, col_offset + i).fill = TOP_FILL
                self.__ws.cell(ROW_START, col_offset + i).border = BORDER 
                self.__ws.cell(ROW_START+1, col_offset + i).border = BORDER 
                self.__ws.cell(DATA_ROW_START, col_offset + i).fill = PHASE_FILL
                self.__ws.cell(DATA_ROW_START, col_offset + i).border = BORDER 

                self.__ws.cell(DATA_ROW_START, col_offset + i).value = date[i].day
            col_offset += len(date)
        return dates

    def __fill_task_date_in_date(self, dates, row, col, card=None, board=None):
        col_offset = col
        for month in dates:
            col_count = 0
            for date in month:
                self.__ws.cell(row, col_offset + col_count).border = BORDER 
                if card != None:
                    if trello_api.is_task_date_in_date(card, date):
                        self.__ws.cell(row, col_offset + col_count).fill = DATE_FILL

                    if trello_api.is_task_actual_date_in_date(card, date):
                        self.__ws.cell(row, col_offset + col_count).value = '○' 

                if board != None:
                    self.__ws.cell(row, col_offset + col_count).fill = PHASE_FILL

                col_count += 1

            col_offset += len(month)


    def performance(self, max_col):
        dates = self.__set_performance_date_cell(max_col)
        row = DATA_ROW_START
        col = max_col + 1
        for board in self.__boards:
            self.__fill_task_date_in_date(dates, row, col, board=board)
            row += 1
            for card in board.get_cards():
                self.__fill_task_date_in_date(dates, row, col, card=card)
                row += 1
                

    def task_ids(self): 
        col_num = self.__col_offset
        self.__set_item_name_cell(30, 'ID')

        row = DATA_ROW_START
        for board in self.__boards:
            self.__ws.cell(row=row, column=col_num).value = board.get_board_id()
            self.__ws.cell(row=row, column=col_num).fill = PHASE_FILL
            self.__ws.cell(row=row, column=col_num).border = BORDER 
            row += 1
            for card in board.get_cards():
                self.__ws.cell(row=row, column=col_num).border = BORDER 
                self.__ws.cell(row=row, column=col_num).value = card.get_card_id() 
                row += 1

    def task_names(self):
        col_num = self.__col_offset
        self.__set_item_name_cell(30, 'タスク名')
        row = DATA_ROW_START
        for board in self.__boards:
            self.__ws.cell(row=row, column=col_num).value = board.get_name()
            self.__ws.cell(row=row, column=col_num).fill = PHASE_FILL
            self.__ws.cell(row=row, column=col_num).border = BORDER 
            row += 1
            for card in board.get_cards():
                self.__ws.cell(row=row, column=col_num).border = BORDER 
                self.__ws.cell(row=row, column=col_num).value = card.get_name()
                row += 1
    
    def task_description(self):
        col_num = self.__col_offset
        self.__set_item_name_cell(20, '説明')
        row = DATA_ROW_START
        for board in self.__boards:
            self.__ws.cell(row=row, column=col_num).value = board.get_description()
            self.__ws.cell(row=row, column=col_num).fill = PHASE_FILL
            self.__ws.cell(row=row, column=col_num).border = BORDER 
            row += 1
            for card in board.get_cards():
                self.__ws.cell(row=row, column=col_num).border = BORDER 
                self.__ws.cell(row=row, column=col_num).value = card.get_desc()
                row += 1

    def task_start_date(self):
        col_num = self.__col_offset
        self.__set_item_name_cell(15, '開始日')
        row = DATA_ROW_START
        for board in self.__boards:
            self.__ws.cell(row=row, column=col_num).fill = PHASE_FILL
            self.__ws.cell(row=row, column=col_num).border = BORDER 
            row += 1
            for card in board.get_cards():
                self.__ws.cell(row=row, column=col_num).border = BORDER 
                if card.get_date() != trello_api.INVALID_DATE_TIME:
                    self.__ws.cell(row=row, column=col_num).value = trello_api.datetime_to_date(card.get_date())
                row += 1

    def task_due_date(self):
        col_num = self.__col_offset
        self.__set_item_name_cell(15, '終了予定日')
        row = DATA_ROW_START
        for board in self.__boards:
            self.__ws.cell(row=row, column=col_num).fill = PHASE_FILL
            self.__ws.cell(row=row, column=col_num).border = BORDER 
            row += 1
            for card in board.get_cards():
                self.__ws.cell(row=row, column=col_num).border = BORDER 
                if card.get_due() != trello_api.INVALID_DATE_TIME:
                    self.__ws.cell(row=row, column=col_num).value = trello_api.datetime_to_date(card.get_due())
                row += 1

    def task_last_activity_date(self):
        col_num = self.__col_offset
        self.__set_item_name_cell(15, '最終更新日')
        row = DATA_ROW_START
        for board in self.__boards:
            self.__ws.cell(row=row, column=col_num).fill = PHASE_FILL
            self.__ws.cell(row=row, column=col_num).border = BORDER 
            row += 1
            for card in board.get_cards():
                self.__ws.cell(row=row, column=col_num).border = BORDER 
                if card.get_date_last_activity() != trello_api.INVALID_DATE_TIME:
                    self.__ws.cell(row=row, column=col_num).value = trello_api.datetime_to_date(card.get_date_last_activity())
                row += 1

    def task_actual_due_date(self):
        col_num = self.__col_offset
        self.__set_item_name_cell(15, '終了日')
        row = DATA_ROW_START
        for board in self.__boards:
            self.__ws.cell(row=row, column=col_num).fill = PHASE_FILL
            self.__ws.cell(row=row, column=col_num).border = BORDER 
            row += 1
            for card in board.get_cards():
                self.__ws.cell(row=row, column=col_num).border = BORDER 
                if card.get_due_complete() or card.get_closed():
                    self.__ws.cell(row=row, column=col_num).value = trello_api.datetime_to_date(card.get_date_last_activity())
                row += 1

    def task_list_name(self):
        col_num = self.__col_offset
        self.__set_item_name_cell(15, 'ステータス')
        row = DATA_ROW_START
        for board in self.__boards:
            self.__ws.cell(row=row, column=col_num).fill = PHASE_FILL
            self.__ws.cell(row=row, column=col_num).border = BORDER 
            row += 1
            for card in board.get_cards():
                self.__ws.cell(row=row, column=col_num).border = BORDER 
                self.__ws.cell(row=row, column=col_num).value = card.get_listname()
                row += 1

    def task_members(self):
        col_num = self.__col_offset
        self.__set_item_name_cell(10, '担当者')
        row = DATA_ROW_START
        for board in self.__boards:
            self.__ws.cell(row=row, column=col_num).fill = PHASE_FILL
            self.__ws.cell(row=row, column=col_num).border = BORDER 
            row += 1
            for card in board.get_cards():
                self.__ws.cell(row=row, column=col_num).border = BORDER 
                self.__ws.cell(row=row, column=col_num).value = ','.join(card.get_membernames())
                row += 1

    def task_indicator(self):
        col_num = self.__col_offset
        self.__set_item_name_cell(10, '進捗率')
        row = DATA_ROW_START
        for board in self.__boards:
            self.__ws.cell(row=row, column=col_num).fill = PHASE_FILL
            self.__ws.cell(row=row, column=col_num).border = BORDER 
            row += 1
            for card in board.get_cards():
                self.__ws.cell(row=row, column=col_num).border = BORDER 
                for check_list in card.get_check_list():
                    self.__ws.cell(row=row, column=col_num).value = trello_api.calc_progress(check_list.get_check_items())
                row += 1

    def filled_task(self):
        row = DATA_ROW_START
        for board in self.__boards:
            row += 1
            for card in board.get_cards():
                if card.get_due_complete() or card.get_closed():
                    self.__fill_all_column('BBBBBB', row)
                elif date.today() >= trello_api.datetime_to_date(card.get_due()):
                    self.__fill_all_column('FF3333', row)
                row += 1

    def header(self):
        self.__ws.cell(HEADER_ROW_START, HEADER_COL_START).value = 'チームID'
        self.__ws.cell(HEADER_ROW_START, HEADER_COL_START).border = BORDER
        self.__ws.cell(HEADER_ROW_START, HEADER_COL_START).fill = PHASE_FILL

        self.__ws.cell(HEADER_ROW_START, HEADER_COL_START + 1).border = BORDER
        self.__ws.cell(HEADER_ROW_START + 1, HEADER_COL_START).value = 'チームメンバー'
        self.__ws.cell(HEADER_ROW_START + 1, HEADER_COL_START).border = BORDER
        self.__ws.cell(HEADER_ROW_START + 1, HEADER_COL_START).fill = PHASE_FILL

        if len(self.__boards) > 0:
            members = ', '.join(str(member) for member in self.__boards[0].get_members().values())

        self.__ws.cell(HEADER_ROW_START + 1, HEADER_COL_START + 1).value = members

        self.__ws.cell(HEADER_ROW_START + 1, HEADER_COL_START + 1).border = BORDER
        self.__ws.cell(HEADER_ROW_START + 2, HEADER_COL_START).value = '作成者'
        self.__ws.cell(HEADER_ROW_START + 2, HEADER_COL_START).border = BORDER
        self.__ws.cell(HEADER_ROW_START + 2, HEADER_COL_START).fill = PHASE_FILL

        self.__ws.cell(HEADER_ROW_START + 2, HEADER_COL_START + 1).border = BORDER

        self.__ws.cell(HEADER_ROW_START, self.__col_offset - 1).value = '実績'
        self.__ws.cell(HEADER_ROW_START, self.__col_offset - 1).border = BORDER
        self.__ws.cell(HEADER_ROW_START, self.__col_offset).value = '○'
        self.__ws.cell(HEADER_ROW_START, self.__col_offset).border = BORDER
        self.__ws.cell(HEADER_ROW_START + 1, self.__col_offset - 1).value = '計画'
        self.__ws.cell(HEADER_ROW_START + 1, self.__col_offset - 1).border = BORDER
        self.__ws.cell(HEADER_ROW_START + 1, self.__col_offset).fill = DATE_FILL
        self.__ws.cell(HEADER_ROW_START + 1, self.__col_offset).border = BORDER 


    def __fill_all_column(self, fill_color, row):
        for col in range(1, self.__ws.max_column + 1):
            self.__ws.cell(row=row, column=col).border = BORDER 
            self.__ws.cell(row, col).fill = PatternFill(fill_type='solid', fgColor=fill_color)

    def exportAsExcel(self):
        if len(self.__boards) < 1:
            return

        self.task_ids() 
        self.task_names()
        self.task_description()
        self.task_members()
        self.task_list_name()
        self.task_indicator()
        self.task_start_date()
        self.task_due_date()
        self.task_actual_due_date()
        self.task_last_activity_date()
        self.__ws.freeze_panes = get_column_letter(self.__col_offset) + str(ROW_START)
        self.performance(self.__ws.max_column)
        self.filled_task()
        self.header()

        today = date.today()
        file_name = date.strftime(today, '%Y-%m-%d')
        print('file exported to : xlsxs/' + file_name + '.xlsx')
        try:
            self.__wb.save('xlsxs/' + file_name + '.xlsx')
        except FileNotFoundError:
            os.mkdir('xlsxs')
            self.__wb.save('xlsxs/' + file_name + '.xlsx')

